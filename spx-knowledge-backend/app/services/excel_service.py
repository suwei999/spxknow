"""
Excel 文档解析服务
支持 .xlsx, .xls, .xlsb, .csv 格式
输出结构与 DOCX/PDF 解析保持一致
"""

import os
import re
import csv
import hashlib
import tempfile
import zipfile
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass

from sqlalchemy.orm import Session
from charset_normalizer import from_bytes

from app.core.logging import logger
from app.config.settings import settings


@dataclass
class ExcelParseOptions:
    """Excel 解析选项"""
    sheet_whitelist: Optional[List[str]] = None
    row_limit_per_sheet: Optional[int] = None
    window_rows: int = 50
    overlap_rows: int = 10
    chunk_max: int = 1000
    chunk_overlap: int = 200


DEFAULT_STYLE_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font></fonts>
  <fills count="1"><fill><patternFill patternType="none"/></fill></fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>
"""


class ExcelService:
    """Excel 文档解析服务，输出结构与 DOCX/PDF 解析保持一致"""

    def __init__(self, db: Session):
        self.db = db
        self.detected_encoding: Optional[str] = None
        self.encoding_confidence: Optional[float] = None
        self.audit_trail: List[Dict[str, Any]] = []

    def parse_document(self, file_path: str, options: Optional[ExcelParseOptions] = None) -> Dict[str, Any]:
        """
        解析 Excel 文档
        
        返回结构：
        - text_content: str - 完整文本内容
        - ordered_elements: List[Dict] - 有序元素列表
        - filtered_elements_light: List[Dict] - 轻量级元素列表
        - text_element_index_map: List[Dict] - 文本索引映射
        - tables: List[Dict] - 表格数据
        - images: List[Dict] - 图片数据
        - metadata: Dict - 元数据（包含 sheet 信息）
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if options is None:
            options = ExcelParseOptions()

        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.csv':
            return self._parse_csv(file_path, options)
        else:
            return self._parse_excel(file_path, options)

    def _parse_csv(self, file_path: str, options: ExcelParseOptions) -> Dict[str, Any]:
        """解析 CSV 文件"""
        logger.info(f"[Excel] 开始解析 CSV 文件: {file_path}")
        
        # 检测编码
        with open(file_path, 'rb') as f:
            raw_bytes = f.read()
        self._detect_encoding(raw_bytes)
        
        # 读取 CSV
        encoding = self.detected_encoding or 'utf-8'
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
        except Exception as e:
            logger.warning(f"[Excel] 使用编码 {encoding} 读取失败，尝试 UTF-8: {e}")
            with open(file_path, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
            self.detected_encoding = 'utf-8'

        if not rows:
            raise ValueError("CSV 文件为空")

        # 检测表头
        header_row_idx = self._detect_header(rows[:10])
        if header_row_idx is not None:
            headers = [str(cell).strip() for cell in rows[header_row_idx]]
            data_rows = rows[header_row_idx + 1:]
        else:
            headers = [f"列{i+1}" for i in range(len(rows[0]))]
            data_rows = rows

        # 限制行数
        if options.row_limit_per_sheet and len(data_rows) > options.row_limit_per_sheet:
            data_rows = data_rows[:options.row_limit_per_sheet]
            self.audit_trail.append({
                "event": "row_limit_applied",
                "sheet": "__csv__",
                "original_rows": len(rows),
                "limited_rows": len(data_rows)
            })

        # 转换为字典列表
        data = []
        for row in data_rows:
            if not any(cell for cell in row):  # 跳过空行
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else ""
                row_dict[header] = self._normalize_cell_value(value)
            data.append(row_dict)

        # 生成 sheet 元数据
        sheet_meta = {
            "name": "__csv__",
            "rows": len(data),
            "columns": len(headers),
            "has_merge": False,
            "has_formula": False,
            "header_detected": header_row_idx is not None,
            "sheet_type": "tabular",
            "layout_features": [],
            "numeric_columns": self._detect_numeric_columns(data, headers),
            "datetime_columns": self._detect_datetime_columns(data, headers),
        }

        # 生成预览样本（默认100行，根据设计文档要求）
        preview_row_limit = 100
        preview_samples = []
        for i, row_dict in enumerate(data[:preview_row_limit]):
            preview_samples.append(row_dict)

        # 生成 chunks
        ordered_elements, filtered_elements, text_elements, tables_meta = self._build_chunks_from_sheet(
            sheet_meta, data, headers, options
        )

        metadata = {
            "element_count": len(ordered_elements),
            "sheet_count": 1,
            "sheets": [sheet_meta],
            "csv_encoding": self.detected_encoding,
            "preview_samples": {"__csv__": preview_samples},
            "row_limit_hit": options.row_limit_per_sheet and len(data_rows) >= options.row_limit_per_sheet,
            "audit_trail": self.audit_trail,
        }

        # 生成完整文本内容
        text_content = self._generate_text_content(sheet_meta, data, headers)

        logger.info(f"[Excel] CSV 解析完成: 行数={len(data)}, 列数={len(headers)}")

        return {
            "text_content": text_content,
            "ordered_elements": ordered_elements,
            "filtered_elements_light": filtered_elements,
            "text_element_index_map": text_elements,
            "tables": [],
            "images": [],
            "tables": tables_meta,
            "metadata": metadata,
        }

    def _parse_excel(self, file_path: str, options: ExcelParseOptions) -> Dict[str, Any]:
        """解析 Excel 文件（.xlsx, .xls, .xlsb）"""
        logger.info(f"[Excel] 开始解析 Excel 文件: {file_path}")

        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("缺少依赖 openpyxl，请先安装: pip install openpyxl")

        file_ext = os.path.splitext(file_path)[1].lower()
        
        # 加载工作簿
        try:
            if file_ext in ('.xlsx', '.xlsm'):
                workbook = self._load_workbook_with_fallback(file_path)
            elif file_ext == '.xls':
                try:
                    import xlrd
                    # xlrd 2.0+ 不支持 .xls，需要 xlrd<2.0 或使用其他库
                    logger.warning("[Excel] .xls 格式需要额外处理，建议转换为 .xlsx")
                    workbook = self._load_workbook_with_fallback(file_path)
                except Exception:
                    raise RuntimeError(".xls 格式暂不支持，请转换为 .xlsx")
            elif file_ext == '.xlsb':
                try:
                    import pyxlsb
                    workbook = pyxlsb.open_workbook(file_path)
                except ImportError:
                    raise RuntimeError("缺少依赖 pyxlsb，请先安装: pip install pyxlsb")
            else:
                raise ValueError(f"不支持的 Excel 格式: {file_ext}")
        except Exception as e:
            logger.error(f"[Excel] 加载工作簿失败: {e}")
            raise

        # 扫描所有 sheet
        sheet_names = workbook.sheetnames
        if options.sheet_whitelist:
            sheet_names = [name for name in sheet_names if name in options.sheet_whitelist]
        
        if len(sheet_names) > 20:
            logger.warning(f"[Excel] Sheet 数量超过 20，仅处理前 20 个")
            sheet_names = sheet_names[:20]
            self.audit_trail.append({
                "event": "sheet_limit_applied",
                "total_sheets": len(workbook.sheetnames),
                "processed_sheets": len(sheet_names)
            })

        all_ordered_elements: List[Dict[str, Any]] = []
        all_filtered_elements: List[Dict[str, Any]] = []
        all_text_elements: List[Dict[str, Any]] = []
        all_images: List[Dict[str, Any]] = []
        sheets_meta: List[Dict[str, Any]] = []
        all_tables: List[Dict[str, Any]] = []
        preview_samples_dict: Dict[str, List[Dict[str, Any]]] = {}

        element_index = 0

        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
                
                # 读取数据
                rows_data = []
                max_row = min(sheet.max_row, options.row_limit_per_sheet or sheet.max_row)
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, values_only=False), start=1):
                    row_values = []
                    for cell in row:
                        value = self._get_cell_value(cell)
                        row_values.append(value)
                    if any(v for v in row_values):  # 跳过全空行
                        rows_data.append(row_values)

                if not rows_data:
                    continue

                # 检测表头
                header_row_idx = self._detect_header_from_rows(rows_data[:10])
                if header_row_idx is not None:
                    headers = [str(cell).strip() if cell else f"列{i+1}" for i, cell in enumerate(rows_data[header_row_idx])]
                    data_rows = rows_data[header_row_idx + 1:]
                else:
                    headers = [f"列{i+1}" for i in range(len(rows_data[0]))]
                    data_rows = rows_data

                # 转换为字典列表
                data = []
                for row in data_rows:
                    row_dict = {}
                    for i, header in enumerate(headers):
                        value = row[i] if i < len(row) else ""
                        row_dict[header] = self._normalize_cell_value(value)
                    data.append(row_dict)

                # 检测 sheet 类型和特征
                has_merge = self._check_merged_cells(sheet)
                has_formula = self._check_formulas(sheet)
                sheet_type, layout_features = self._detect_sheet_type(sheet, data, headers)

                # 提取图片
                sheet_images = self._extract_images(sheet, sheet_name, element_index)
                all_images.extend(sheet_images)

                # 生成 sheet 元数据
                sheet_meta = {
                    "name": sheet_name,
                    "rows": len(data),
                    "columns": len(headers),
                    "has_merge": has_merge,
                    "has_formula": has_formula,
                    "header_detected": header_row_idx is not None,
                    "sheet_type": sheet_type,
                    "layout_features": layout_features,
                    "numeric_columns": self._detect_numeric_columns(data, headers),
                    "datetime_columns": self._detect_datetime_columns(data, headers),
                }
                sheets_meta.append(sheet_meta)

                # 生成预览样本
                preview_samples = []
                for i, row_dict in enumerate(data[:3]):
                    preview_samples.append(row_dict)
                preview_samples_dict[sheet_name] = preview_samples

                # 生成 chunks
                sheet_elements, sheet_filtered, sheet_text, sheet_tables = self._build_chunks_from_sheet(
                    sheet_meta, data, headers, options, element_index
                )
                
                all_ordered_elements.extend(sheet_elements)
                all_filtered_elements.extend(sheet_filtered)
                all_text_elements.extend(sheet_text)
                all_tables.extend(sheet_tables)
                element_index += len(sheet_elements)

            except Exception as e:
                logger.error(f"[Excel] 处理 Sheet '{sheet_name}' 失败: {e}", exc_info=True)
                self.audit_trail.append({
                    "event": "sheet_parse_error",
                    "sheet": sheet_name,
                    "error": str(e)
                })
                continue

        # 关闭工作簿
        if hasattr(workbook, 'close'):
            workbook.close()

        metadata = {
            "element_count": len(all_ordered_elements),
            "sheet_count": len(sheets_meta),
            "sheets": sheets_meta,
            "preview_samples": preview_samples_dict,
            "row_limit_hit": any(
                options.row_limit_per_sheet and sheet['rows'] >= options.row_limit_per_sheet
                for sheet in sheets_meta
            ),
            "embedded_objects": [{"type": "image", "sheet": img.get("sheet_name"), "object_url": img.get("object_url")} 
                                for img in all_images],
            "audit_trail": self.audit_trail,
        }

        # 生成完整文本内容
        text_content = self._generate_text_content_from_sheets(sheets_meta, preview_samples_dict)

        logger.info(f"[Excel] Excel 解析完成: Sheet数={len(sheets_meta)}, 元素数={len(all_ordered_elements)}")

        return {
            "text_content": text_content,
            "ordered_elements": all_ordered_elements,
            "filtered_elements_light": all_filtered_elements,
            "text_element_index_map": all_text_elements,
            "tables": all_tables,
            "images": all_images,
            "metadata": metadata,
        }

    def _detect_encoding(self, raw_bytes: bytes) -> None:
        """检测文件编码（用于 CSV）"""
        match = from_bytes(raw_bytes).best()
        if match:
            self.detected_encoding = match.encoding
            if hasattr(match, 'confidence'):
                self.encoding_confidence = match.confidence
            elif hasattr(match, 'percent_coherence'):
                self.encoding_confidence = match.percent_coherence / 100.0
            elif hasattr(match, 'coherence'):
                self.encoding_confidence = match.coherence
            else:
                self.encoding_confidence = None
        else:
            self.detected_encoding = "utf-8"
            self.encoding_confidence = None

    def _detect_header(self, rows: List[List[str]], max_check: int = 5) -> Optional[int]:
        """检测表头行（CSV）"""
        if not rows:
            return None
        
        best_score = 0
        best_idx = None
        
        for idx in range(min(len(rows), max_check)):
            row = rows[idx]
            if not row:
                continue
            
            # 计算非空单元格比例
            non_empty = sum(1 for cell in row if cell and str(cell).strip())
            if non_empty == 0:
                continue
            
            ratio = non_empty / len(row) if row else 0
            
            # 检查是否包含文本特征（非纯数字）
            text_count = sum(1 for cell in row if cell and not str(cell).strip().replace('.', '').replace('-', '').isdigit())
            text_ratio = text_count / len(row) if row else 0
            
            score = ratio * 0.6 + text_ratio * 0.4
            if score > best_score:
                best_score = score
                best_idx = idx
        
        return best_idx if best_score > 0.5 else None

    def _detect_header_from_rows(self, rows: List[List[Any]]) -> Optional[int]:
        """检测表头行（Excel）"""
        return self._detect_header([[str(cell) if cell is not None else "" for cell in row] for row in rows])

    def _get_cell_value(self, cell) -> Any:
        """获取单元格值（处理公式、日期等）"""
        if cell is None:
            return None
        
        try:
            import openpyxl
            if isinstance(cell, openpyxl.cell.cell.Cell):
                # 优先使用计算后的值
                if cell.data_type == 'f':  # 公式
                    try:
                        return cell.value  # data_only=True 时已计算
                    except:
                        return f"={cell.value}" if cell.value else None
                elif cell.data_type == 'd':  # 日期
                    if isinstance(cell.value, datetime):
                        return cell.value.isoformat()
                    return cell.value
                else:
                    return cell.value
        except:
            pass
        
        return cell.value if hasattr(cell, 'value') else str(cell)

    def _normalize_cell_value(self, value: Any) -> str:
        """归一化单元格值为字符串"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

    def _check_merged_cells(self, sheet) -> bool:
        """检查是否有合并单元格"""
        try:
            if hasattr(sheet, 'merged_cells'):
                return len(list(sheet.merged_cells)) > 0
        except:
            pass
        return False

    def _check_formulas(self, sheet) -> bool:
        """检查是否有公式"""
        try:
            import openpyxl
            if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                for row in sheet.iter_rows(max_row=min(100, sheet.max_row)):
                    for cell in row:
                        if cell.data_type == 'f':
                            return True
        except:
            pass
        return False

    def _detect_sheet_type(self, sheet, data: List[Dict], headers: List[str]) -> Tuple[str, List[str]]:
        """检测 sheet 类型（tabular 或 layout）"""
        layout_features = []
        
        # 检查合并单元格
        if self._check_merged_cells(sheet):
            layout_features.append("merged_cells")
        
        # 检查是否有图表/图片
        try:
            if hasattr(sheet, '_images') and sheet._images:
                layout_features.append("images")
            if hasattr(sheet, '_charts') and sheet._charts:
                layout_features.append("charts")
        except:
            pass
        
        # 检查列宽/行高是否异常（可能是布局型）
        try:
            if hasattr(sheet, 'column_dimensions'):
                col_widths = [sheet.column_dimensions.get(chr(65+i)).width for i in range(min(10, len(headers)))]
                if any(w and w > 50 for w in col_widths if w):
                    layout_features.append("wide_columns")
        except:
            pass
        
        # 如果数据行数很少但列很多，可能是布局型
        if len(data) < 10 and len(headers) > 5:
            layout_features.append("sparse_data")
        
        if layout_features:
            return "layout", layout_features
        return "tabular", []

    def _extract_images(self, sheet, sheet_name: str, base_element_index: int) -> List[Dict[str, Any]]:
        """提取 sheet 中的图片"""
        images = []
        try:
            import openpyxl
            if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                if hasattr(sheet, '_images') and sheet._images:
                    for idx, img in enumerate(sheet._images):
                        try:
                            # 获取图片数据
                            img_data = img._data()
                            if img_data:
                                images.append({
                                    "data": img_data,
                                    "element_index": base_element_index + idx,
                                    "sheet_name": sheet_name,
                                    "image_format": img.format,
                                    "anchor": str(img.anchor) if hasattr(img, 'anchor') else None,
                                })
                        except Exception as e:
                            logger.warning(f"[Excel] 提取图片失败: {e}")
        except Exception as e:
            logger.debug(f"[Excel] 检查图片失败: {e}")
        
        return images

    def _detect_numeric_columns(self, data: List[Dict], headers: List[str]) -> List[str]:
        """检测数值列"""
        numeric_cols = []
        if not data:
            return numeric_cols
        
        for header in headers:
            numeric_count = 0
            for row in data[:100]:  # 采样前100行
                value = row.get(header, "")
                if value and isinstance(value, (int, float)):
                    numeric_count += 1
                elif value and str(value).replace('.', '').replace('-', '').isdigit():
                    numeric_count += 1
            
            if numeric_count > len(data) * 0.5:  # 超过50%是数字
                numeric_cols.append(header)
        
        return numeric_cols

    def _detect_datetime_columns(self, data: List[Dict], headers: List[str]) -> List[str]:
        """检测日期列"""
        datetime_cols = []
        if not data:
            return datetime_cols
        
        for header in headers:
            datetime_count = 0
            for row in data[:100]:  # 采样前100行
                value = row.get(header, "")
                if isinstance(value, datetime):
                    datetime_count += 1
                elif value and isinstance(value, str):
                    # 尝试解析常见日期格式
                    try:
                        datetime.fromisoformat(value.replace('Z', '+00:00'))
                        datetime_count += 1
                    except:
                        pass
            
            if datetime_count > len(data) * 0.3:  # 超过30%是日期
                datetime_cols.append(header)
        
        return datetime_cols

    def _build_chunks_from_sheet(
        self,
        sheet_meta: Dict[str, Any],
        data: List[Dict],
        headers: List[str],
        options: ExcelParseOptions,
        base_element_index: int = 0
    ) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
        """从 sheet 数据构建 chunks"""
        ordered_elements: List[Dict] = []
        filtered_elements: List[Dict] = []
        text_elements: List[Dict] = []
        tables_meta: List[Dict] = []
        
        sheet_name = sheet_meta["name"]
        sheet_type = sheet_meta.get("sheet_type", "tabular")
        
        if sheet_type == "tabular" and data:
            # 生成 tabular chunks
            window_rows = options.window_rows
            overlap_rows = options.overlap_rows
            
            for start_idx in range(0, len(data), window_rows - overlap_rows):
                end_idx = min(start_idx + window_rows, len(data))
                window_data = data[start_idx:end_idx]
                
                # 检查是否需要降采样列
                columns_to_keep = headers
                columns_dropped = []
                if len(headers) > 20:  # 宽表降采样
                    # 优先保留非空率高的列
                    col_scores = {}
                    for header in headers:
                        non_empty = sum(1 for row in window_data if row.get(header))
                        col_scores[header] = non_empty / len(window_data) if window_data else 0
                    
                    # 排序并保留前10列 + 数值列
                    sorted_cols = sorted(col_scores.items(), key=lambda x: x[1], reverse=True)
                    columns_to_keep = [col[0] for col in sorted_cols[:10]]
                    columns_to_keep.extend(sheet_meta.get("numeric_columns", []))
                    columns_to_keep = list(dict.fromkeys(columns_to_keep))  # 去重
                    columns_dropped = [h for h in headers if h not in columns_to_keep]
                    
                    self.audit_trail.append({
                        "event": "column_downsample",
                        "sheet": sheet_name,
                        "window": f"{start_idx}-{end_idx}",
                        "original_columns": len(headers),
                        "kept_columns": len(columns_to_keep),
                        "dropped_columns": columns_dropped
                    })
                
                # 生成 Markdown 表格
                markdown_table = self._generate_markdown_table(headers, window_data, columns_to_keep)
                
                # 计算统计信息
                summary_stats = {}
                for col in sheet_meta.get("numeric_columns", []):
                    if col in columns_to_keep:
                        values = [float(row.get(col, 0)) for row in window_data if row.get(col)]
                        if values:
                            summary_stats[col] = {
                                "min": min(values),
                                "max": max(values),
                                "avg": sum(values) / len(values),
                                "count": len(values)
                            }
                
                element_index = base_element_index + len(ordered_elements)
                
                table_cells = [
                    [str(row.get(col, "")) for col in columns_to_keep]
                    for row in window_data
                ]
                table_data = {
                    "cells": table_cells,
                    "rows": len(table_cells),
                    "columns": len(columns_to_keep),
                    "headers": columns_to_keep,
                    "row_start": start_idx + 1,
                    "row_end": end_idx,
                    "structure": "excel_tabular_window",
                    "sheet_name": sheet_name,
                }
                chunk_meta = {
                    "sheet_name": sheet_name,
                    "sheet_type": sheet_type,
                    "row_start": start_idx + 1,
                    "row_end": end_idx,
                    "column_headers": columns_to_keep,
                    "chunk_type": "tabular",
                }
                
                if columns_dropped:
                    chunk_meta["columns_dropped"] = columns_dropped
                if summary_stats:
                    chunk_meta["summary_stats"] = summary_stats
                
                element = {
                    "type": "table",
                    "text": markdown_table,
                    "element_index": element_index,
                    "doc_order": element_index,
                    "chunk_type": "tabular",
                    "table_data": table_data,
                    "table_text": markdown_table,
                    "metadata": chunk_meta,
                }
                
                tables_meta.append({
                    "element_index": element_index,
                    "table_data": table_data,
                    "table_text": markdown_table,
                    "sheet_name": sheet_name,
                    "doc_order": element_index,
                })
                ordered_elements.append(element)
                filtered_elements.append({
                    "category": "Table",
                    "text": markdown_table,
                    "element_index": element_index,
                    "chunk_type": "tabular",
                })
                text_elements.append({
                    "element_index": element_index,
                    "element_type": "table",
                    "chunk_type": "tabular",
                    "sheet_name": sheet_name,
                })
        
        # 生成文本 chunks（滑动窗口）
        enable_flatten = getattr(settings, "EXCEL_ENABLE_FLATTENED_TEXT", False)
        should_flatten = enable_flatten or sheet_type != "tabular"
        if should_flatten:
            full_text = self._generate_text_content(sheet_meta, data, headers)
            text_chunks = self._split_text_chunks(full_text, options.chunk_max, options.chunk_overlap)
            
            for idx, chunk_text in enumerate(text_chunks):
                element_index = base_element_index + len(ordered_elements)
                
                element = {
                    "type": "text",
                    "text": chunk_text,
                    "element_index": element_index,
                    "doc_order": element_index,
                    "chunk_type": "text",
                    "metadata": {
                        "sheet_name": sheet_name,
                        "sheet_type": sheet_type,
                        "chunk_type": "text",
                    },
                }
                
                ordered_elements.append(element)
                filtered_elements.append({
                    "category": "text",
                    "text": chunk_text,
                    "element_index": element_index,
                    "chunk_type": "text",
                })
                text_elements.append({
                    "element_index": element_index,
                    "element_type": "text",
                    "chunk_type": "text",
                    "sheet_name": sheet_name,
                })
        
        return ordered_elements, filtered_elements, text_elements, tables_meta

    def _generate_markdown_table(self, headers: List[str], data: List[Dict], columns_to_keep: List[str]) -> str:
        """生成 Markdown 表格"""
        if not headers or not data:
            return ""
        
        # 过滤列
        visible_headers = [h for h in headers if h in columns_to_keep]
        if not visible_headers:
            visible_headers = headers[:10]  # 至少显示前10列
        
        lines = []
        # 表头
        lines.append("| " + " | ".join(visible_headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(visible_headers)) + " |")
        
        # 数据行
        for row in data:
            cells = [str(row.get(h, ""))[:100] for h in visible_headers]  # 限制单元格长度
            lines.append("| " + " | ".join(cells) + " |")
        
        return "\n".join(lines)

    def _generate_text_content(self, sheet_meta: Dict, data: List[Dict], headers: List[str]) -> str:
        """生成完整文本内容"""
        lines = [f"Sheet: {sheet_meta['name']}"]
        if headers:
            lines.append(f"列: {', '.join(headers)}")
        lines.append("")
        
        for row in data[:100]:  # 限制行数
            row_text = " | ".join([f"{h}: {row.get(h, '')}" for h in headers[:10]])
            lines.append(row_text)
        
        return "\n".join(lines)

    def _generate_text_content_from_sheets(
        self,
        sheets_meta: List[Dict],
        preview_samples: Dict[str, List[Dict]]
    ) -> str:
        """从多个 sheet 生成完整文本内容"""
        parts = []
        for sheet_meta in sheets_meta:
            sheet_name = sheet_meta["name"]
            samples = preview_samples.get(sheet_name, [])
            part = self._generate_text_content(sheet_meta, samples, list(samples[0].keys()) if samples else [])
            parts.append(part)
        return "\n\n".join(parts)

    def _split_text_chunks(self, text: str, chunk_max: int, chunk_overlap: int) -> List[str]:
        """将文本按滑动窗口分块"""
        if len(text) <= chunk_max:
            return [text] if text.strip() else []
        
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_max
            chunk = text[start:end]
            if chunk.strip():
                chunks.append(chunk.strip())
            start = end - chunk_overlap
        
        return chunks

    def _load_workbook_with_fallback(self, file_path: str):
        """加载工作簿，遇到样式异常时自动降级"""
        import openpyxl  # 局部导入，避免模块级依赖
        try:
            return openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except TypeError as exc:
            logger.warning("[Excel] openpyxl 解析样式失败，尝试移除样式后重试: %s", exc)
            self.audit_trail.append({
                "event": "styles_fallback",
                "message": "openpyxl 无法解析样式，已使用无样式副本",
            })
            sanitized_bytes = self._strip_styles(file_path)
            bio = BytesIO(sanitized_bytes)
            return openpyxl.load_workbook(bio, data_only=True, read_only=True)

    def _strip_styles(self, file_path: str) -> bytes:
        """将 styles.xml 替换为最小样式，返回新的 xlsx 字节"""
        output = BytesIO()
        replaced = False
        with zipfile.ZipFile(file_path, 'r') as zin, zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    zout.writestr(item, DEFAULT_STYLE_XML.encode('utf-8'))
                    replaced = True
                else:
                    zout.writestr(item, data)
        if not replaced:
            # 没有样式文件，直接返回原始内容
            with open(file_path, 'rb') as f:
                return f.read()
        return output.getvalue()
